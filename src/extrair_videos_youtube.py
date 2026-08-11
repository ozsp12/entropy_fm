# -*- coding: utf-8 -*-
"""Extract Entropy FM YouTube metadata to an Excel workbook."""

import argparse
import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

CANAL_HANDLE = "@entropyfm"
ENDERECO_DO_CANAL = "https://www.youtube.com/@entropyfm"
ARQUIVO_DE_SAIDA = "data/videos_entropy_fm.xlsx"


def converter_duracao_iso(texto):
    padrao = re.match(
        r"P(?:(\d+)D)?T(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", texto or ""
    )
    if not padrao:
        return ""
    dias, horas, minutos, segundos = (int(g or 0) for g in padrao.groups())
    horas += dias * 24
    return "{:02d}:{:02d}:{:02d}".format(horas, minutos, segundos)


def resolver_canal(servico):
    resposta = servico.channels().list(
        part="id,contentDetails,snippet",
        forHandle=CANAL_HANDLE,
    ).execute()

    itens = resposta.get("items", [])
    if not itens:
        raise RuntimeError("Canal Entropy FM não encontrado pela YouTube Data API.")
    return itens[0]


def coletar_videos(chave):
    from googleapiclient.discovery import build

    servico = build("youtube", "v3", developerKey=chave)
    canal = resolver_canal(servico)
    identificador_do_canal = canal["id"]
    lista_de_envios = canal["contentDetails"]["relatedPlaylists"]["uploads"]

    registros = []
    token_da_pagina = None

    while True:
        resposta = servico.playlistItems().list(
            part="snippet,contentDetails",
            playlistId=lista_de_envios,
            maxResults=50,
            pageToken=token_da_pagina,
        ).execute()

        for item in resposta.get("items", []):
            resumo = item["snippet"]
            identificador = item["contentDetails"]["videoId"]
            registros.append({
                "identificador": identificador,
                "titulo": resumo.get("title", ""),
                "descricao": resumo.get("description", ""),
                "publicacao": item["contentDetails"].get("videoPublishedAt", ""),
                "link": "https://www.youtube.com/watch?v=" + identificador,
                "duracao": "",
            })

        token_da_pagina = resposta.get("nextPageToken")
        if not token_da_pagina:
            break

    indice = {registro["identificador"]: registro for registro in registros}
    ids = list(indice.keys())
    for inicio in range(0, len(ids), 50):
        bloco = ids[inicio:inicio + 50]
        resposta = servico.videos().list(
            part="contentDetails",
            id=",".join(bloco),
        ).execute()
        for item in resposta.get("items", []):
            indice[item["id"]]["duracao"] = converter_duracao_iso(
                item["contentDetails"]["duration"]
            )

    registros.sort(key=lambda registro: registro["publicacao"], reverse=True)
    return registros, identificador_do_canal


def gravar_planilha(registros, identificador_do_canal, caminho):
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    cabecalhos = ["Número", "Título", "Descrição", "Link", "Publicação", "Duração"]
    larguras = [8, 60, 90, 45, 14, 12]

    livro = Workbook()
    planilha = livro.active
    planilha.title = "Videos"

    for coluna, (rotulo, largura) in enumerate(zip(cabecalhos, larguras), start=1):
        celula = planilha.cell(row=1, column=coluna, value=rotulo)
        celula.font = Font(name="Arial", bold=True)
        celula.alignment = Alignment(vertical="center")
        planilha.column_dimensions[get_column_letter(coluna)].width = largura

    for linha, registro in enumerate(registros, start=2):
        valores = [
            linha - 1,
            registro["titulo"],
            registro["descricao"],
            registro["link"],
            registro["publicacao"][:10],
            registro["duracao"],
        ]
        for coluna, valor in enumerate(valores, start=1):
            celula = planilha.cell(row=linha, column=coluna, value=valor)
            celula.font = Font(name="Arial")
            celula.alignment = Alignment(vertical="top", wrap_text=(coluna == 3))

    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = "A1:F{}".format(max(len(registros) + 1, 2))

    nota = livro.create_sheet("Notas")
    nota["A1"] = "Canal"
    nota["B1"] = ENDERECO_DO_CANAL
    nota["A2"] = "Handle"
    nota["B2"] = CANAL_HANDLE
    nota["A3"] = "Identificador do canal"
    nota["B3"] = identificador_do_canal
    nota["A4"] = "Extração"
    nota["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    nota["A5"] = "Total de vídeos"
    nota["B5"] = len(registros)

    for linha in range(1, 6):
        nota.cell(row=linha, column=1).font = Font(name="Arial", bold=True)
        nota.cell(row=linha, column=2).font = Font(name="Arial")

    nota.column_dimensions["A"].width = 26
    nota.column_dimensions["B"].width = 55

    livro.save(caminho)


def principal():
    analisador = argparse.ArgumentParser()
    analisador.add_argument(
        "--chave",
        default=os.environ.get("YOUTUBE_API_KEY"),
        help="YouTube Data API key; defaults to YOUTUBE_API_KEY.",
    )
    analisador.add_argument("--saida", default=ARQUIVO_DE_SAIDA)
    argumentos = analisador.parse_args()

    if not argumentos.chave:
        analisador.error("configure --chave ou a variável YOUTUBE_API_KEY")

    registros, identificador_do_canal = coletar_videos(argumentos.chave)
    gravar_planilha(registros, identificador_do_canal, argumentos.saida)
    print("{} vídeos gravados em {}".format(len(registros), argumentos.saida))


if __name__ == "__main__":
    principal()
